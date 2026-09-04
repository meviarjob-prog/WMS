import io
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .categorize import classify_by_name

NOMENCLATURE_HEADERS = [
    "Артикул (SKU)",
    "Штрихкод",
    "Наименование",
    "Ед. изм.",
    "Описание",
    "Норма времени на 1 шт, мин",
]


def _style_header(ws, headers):
    fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    for idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx, value=title)
        cell.font = Font(bold=True)
        cell.fill = fill
        ws.column_dimensions[get_column_letter(idx)].width = max(16, len(title) + 4)


def build_nomenclature_template() -> bytes:
    """Готовит xlsx-шаблон для заполнения номенклатуры."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Номенклатура"
    _style_header(ws, NOMENCLATURE_HEADERS)

    example = ["ART-0001", "4600000000015", "Пример: Футболка белая XL", "шт", "", 12]
    ws.append(example)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


class ImportResult:
    def __init__(self):
        self.created = 0
        self.updated = 0
        self.errors = []  # list[str]

    @property
    def total(self):
        return self.created + self.updated


def import_nomenclature_from_excel(file_stream, db, Nomenclature) -> ImportResult:
    """Импортирует/обновляет номенклатуру из xlsx-файла (по шаблону)."""
    result = ImportResult()

    try:
        wb = load_workbook(file_stream, data_only=True)
    except Exception as exc:  # noqa: BLE001
        result.errors.append(f"Не удалось открыть файл: {exc}")
        return result

    ws = wb.active

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(v is None or str(v).strip() == "" for v in row):
            continue

        sku = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ""
        barcode = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        name = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
        unit = str(row[3]).strip() if len(row) > 3 and row[3] not in (None, "") else "шт"
        description = str(row[4]).strip() if len(row) > 4 and row[4] is not None else ""
        norm_minutes = None
        if len(row) > 5 and row[5] not in (None, ""):
            try:
                norm_minutes = float(row[5])
            except (TypeError, ValueError):
                result.errors.append(f"Строка {row_idx}: некорректная норма времени '{row[5]}'")

        if not sku or not name:
            result.errors.append(f"Строка {row_idx}: не заполнен артикул или наименование")
            continue

        if not barcode:
            barcode = sku

        existing = Nomenclature.query.filter_by(sku=sku).first()

        barcode_owner = Nomenclature.query.filter_by(barcode=barcode).first()
        if barcode_owner is not None and (existing is None or barcode_owner.id != existing.id):
            result.errors.append(
                f"Строка {row_idx}: штрихкод '{barcode}' уже используется другим товаром"
            )
            continue

        if existing:
            existing.barcode = barcode
            existing.name = name
            existing.unit = unit or "шт"
            existing.description = description
            if norm_minutes is not None:
                existing.norm_minutes = norm_minutes
            # Вид товара переопределять не трогаем, если он уже выставлен
            # (вручную или предыдущим импортом) — только если пуст.
            if existing.category_id is None:
                category = classify_by_name(name)
                existing.category_id = category.id if category else None
            result.updated += 1
        else:
            category = classify_by_name(name)
            item = Nomenclature(
                sku=sku,
                barcode=barcode,
                name=name,
                unit=unit or "шт",
                description=description,
                norm_minutes=norm_minutes,
                category_id=category.id if category else None,
            )
            db.session.add(item)
            result.created += 1

    return result


def export_nomenclature_to_excel(items) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Номенклатура"
    _style_header(ws, NOMENCLATURE_HEADERS)
    for item in items:
        ws.append(
            [item.sku, item.barcode, item.name, item.unit, item.description or "", item.norm_minutes or ""]
        )
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


RECEIVING_HEADERS = [
    "Номер документа",
    "Дата",
    "Склад",
    "Поставщик",
    "Статус",
    "Артикул",
    "Штрихкод",
    "Наименование",
    "Кол-во",
    "Ед. изм.",
]


def export_receiving_to_excel(documents) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Приемка"
    _style_header(ws, RECEIVING_HEADERS)

    status_map = {"draft": "Черновик", "completed": "Завершен"}

    for doc in documents:
        for line in doc.lines:
            ws.append(
                [
                    doc.number,
                    doc.created_at.strftime("%Y-%m-%d %H:%M") if doc.created_at else "",
                    doc.warehouse.name if doc.warehouse else "",
                    doc.supplier or "",
                    status_map.get(doc.status, doc.status),
                    line.nomenclature.sku if line.nomenclature else "",
                    line.nomenclature.barcode if line.nomenclature else "",
                    line.nomenclature.name if line.nomenclature else "",
                    line.qty,
                    line.nomenclature.unit if line.nomenclature else "",
                ]
            )

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


PLACEMENT_HEADERS = [
    "Номер документа",
    "Дата",
    "Склад",
    "Статус",
    "Артикул",
    "Штрихкод",
    "Наименование",
    "Кол-во",
    "Ед. изм.",
    "Короб",
    "Ячейка",
]


def export_placement_to_excel(documents) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Размещение"
    _style_header(ws, PLACEMENT_HEADERS)

    status_map = {"draft": "Черновик", "completed": "Завершен"}

    for doc in documents:
        for line in doc.lines:
            ws.append(
                [
                    doc.number,
                    doc.created_at.strftime("%Y-%m-%d %H:%M") if doc.created_at else "",
                    doc.warehouse.name if doc.warehouse else "",
                    status_map.get(doc.status, doc.status),
                    line.nomenclature.sku if line.nomenclature else "",
                    line.nomenclature.barcode if line.nomenclature else "",
                    line.nomenclature.name if line.nomenclature else "",
                    line.qty,
                    line.nomenclature.unit if line.nomenclature else "",
                    line.box.box_number if line.box else "",
                    line.box.cell.code if (line.box and line.box.cell) else "",
                ]
            )

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


MOVEMENT_HEADERS = [
    "Номер документа",
    "Дата",
    "Статус",
    "Короб",
    "Артикул",
    "Штрихкод",
    "Наименование",
    "Кол-во",
    "Ед. изм.",
    "Склад-источник",
    "Ячейка-источник",
    "Склад-назначение",
    "Ячейка-назначение",
]


def export_movement_to_excel(documents) -> bytes:
    """Одна строка на каждый товар в каждом коробе документа перемещения —
    короб сканируется целиком, но в отчете видно содержимое."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Перемещения"
    _style_header(ws, MOVEMENT_HEADERS)

    status_map = {"draft": "Черновик", "completed": "Завершен"}

    for doc in documents:
        for line in doc.lines:
            box_items = list(line.box.items) if line.box else []
            rows = box_items or [None]
            for box_item in rows:
                ws.append(
                    [
                        doc.number,
                        doc.created_at.strftime("%Y-%m-%d %H:%M") if doc.created_at else "",
                        status_map.get(doc.status, doc.status),
                        line.box.box_number if line.box else "",
                        box_item.nomenclature.sku if box_item else "",
                        box_item.nomenclature.barcode if box_item else "",
                        box_item.nomenclature.name if box_item else "",
                        box_item.qty if box_item else "",
                        box_item.nomenclature.unit if box_item else "",
                        line.from_warehouse.name if line.from_warehouse else "",
                        line.from_cell.code if line.from_cell else "",
                        doc.to_warehouse.name if doc.to_warehouse else "",
                        line.to_cell.code if line.to_cell else "",
                    ]
                )

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


PRODUCTION_HEADERS = [
    "Дата",
    "Сотрудник",
    "Кол-во, шт",
    "Нормо-минуты",
    "Плановая смена, мин",
    "Эффективность, %",
    "Из них без нормы, шт",
]


def export_production_to_excel(rows) -> bytes:
    """rows — список словарей из production._efficiency_rows()."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Эффективность"
    _style_header(ws, PRODUCTION_HEADERS)

    for row in rows:
        ws.append(
            [
                row["work_date"].strftime("%Y-%m-%d") if row["work_date"] else "",
                row["user"].display_name() if row["user"] else "",
                row["qty"],
                row["normo_minutes"],
                row["shift_minutes"],
                row["efficiency"] if row["efficiency"] is not None else "",
                row["missing_norm"],
            ]
        )

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


INVENTORY_HEADERS = [
    "Номер документа",
    "Дата",
    "Склад",
    "Статус",
    "Артикул",
    "Штрихкод",
    "Наименование",
    "Кол-во",
    "Ед. изм.",
]


def export_inventory_to_excel(documents) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Инвентаризация"
    _style_header(ws, INVENTORY_HEADERS)

    status_map = {"draft": "Черновик", "completed": "Завершен"}

    for doc in documents:
        for line in doc.lines:
            ws.append(
                [
                    doc.number,
                    doc.created_at.strftime("%Y-%m-%d %H:%M") if doc.created_at else "",
                    doc.warehouse.name if doc.warehouse else "",
                    status_map.get(doc.status, doc.status),
                    line.nomenclature.sku if line.nomenclature else "",
                    line.nomenclature.barcode if line.nomenclature else "",
                    line.nomenclature.name if line.nomenclature else "",
                    line.qty,
                    line.nomenclature.unit if line.nomenclature else "",
                ]
            )

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def timestamp_for_filename() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")
